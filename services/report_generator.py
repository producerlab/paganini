import json
import os
import re
import asyncio
import pandas as pd
import httpx
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import date, timedelta, datetime
from typing import Any, Dict, List
from aiogram.types import Message
from sqlalchemy.ext.asyncio import AsyncSession

from database.models import Report
from services.logging import logger


# ------------------ Custom Exceptions ------------------
class ReportError(Exception):
    """Base exception for report generation errors"""
    pass


class InvalidTokenError(ReportError):
    """Token is invalid or lacks permissions"""
    pass


class WBTimeoutError(ReportError):
    """WB API timeout"""
    pass


class NoDataError(ReportError):
    """No data found for the period"""
    pass


# ------------------ Progress Stages ------------------
PROGRESS_STAGES = {
    'init': '‚è≥ –ü–æ–¥–≥–æ—Ç–æ–≤–∫–∞ –∫ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏...',
    'fetch_sales': 'üì° –ü–æ–ª—É—á–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö –æ –ø—Ä–æ–¥–∞–∂–∞—Ö –∏–∑ WB...',
    'fetch_ads': 'üìä –ó–∞–≥—Ä—É–∑–∫–∞ –¥–∞–Ω–Ω—ã—Ö –æ —Ä–µ–∫–ª–∞–º–µ...',
    'fetch_cards': 'üè∑ –ó–∞–≥—Ä—É–∑–∫–∞ –∫–∞—Ä—Ç–æ—á–µ–∫ —Ç–æ–≤–∞—Ä–æ–≤...',
    'process': 'üìà –û–±—Ä–∞–±–æ—Ç–∫–∞ –∏ –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏–µ –¥–∞–Ω–Ω—ã—Ö...',
    'create_excel': 'üìÑ –§–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏–µ Excel —Ñ–∞–π–ª–∞...',
}


# ------------------ HTTP‚Äëclients ------------------
SYNC_CLIENT  = httpx.Client(timeout=120.0)
ASYNC_CLIENT = httpx.AsyncClient(timeout=120.0)


async def close_http_clients():
    """Close HTTP clients on shutdown"""
    SYNC_CLIENT.close()
    await ASYNC_CLIENT.aclose()
    logger.info("HTTP clients closed")


async def run_with_progress(message: Message, title: str, coro, progress_state: dict, *args):
    """
    –û—Ç–æ–±—Ä–∞–∂–∞–µ—Ç —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –ø—Ä–æ–≥—Ä–µ—Å—Å–æ–º, –ø–æ–∫–∞ –≤—ã–ø–æ–ª–Ω—è–µ—Ç—Å—è coroutine coro.
    –ö–∞–∂–¥—É—é —Å–µ–∫—É–Ω–¥—É –æ–±–Ω–æ–≤–ª—è–µ—Ç —Ç–µ–∫—Å—Ç —Å–æ–æ–±—â–µ–Ω–∏—è —Å –∏–Ω–¥–∏–∫–∞—Ü–∏–µ–π –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è.
    –ü–æ—Å–ª–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è —Ä–∞–±–æ—Ç—ã coroutine —Å–æ–æ–±—â–µ–Ω–∏–µ —É–¥–∞–ª—è–µ—Ç—Å—è, –∞ —Ä–µ–∑—É–ª—å—Ç–∞—Ç –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç—Å—è.
    –í—ã–±—Ä–∞—Å—ã–≤–∞–µ—Ç —Å–ø–µ—Ü–∏—Ñ–∏—á–Ω—ã–µ –∏—Å–∫–ª—é—á–µ–Ω–∏—è: WBTimeoutError, InvalidTokenError.

    Args:
        message: Telegram message object
        title: Initial progress title
        coro: Coroutine to execute
        progress_state: Dict for sharing progress stage between coroutines
        *args: Arguments for the coroutine
    """
    progress_state['stage'] = 'init'
    current_text = PROGRESS_STAGES.get('init', title)
    progress_message = await message.answer(current_text)
    task = asyncio.create_task(coro(progress_state, *args))
    dots = ['.', '..', '...']
    i = 0
    last_stage = 'init'

    try:
        while not task.done():
            current_stage = progress_state.get('stage', 'init')
            stage_text = PROGRESS_STAGES.get(current_stage, title)
            dot = dots[i % len(dots)]

            # Update message only if stage changed or every second for dots
            try:
                if current_stage != last_stage:
                    await progress_message.edit_text(f"{stage_text}")
                    last_stage = current_stage
                else:
                    await progress_message.edit_text(f"{stage_text}{dot}")
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞: {e}")

            await asyncio.sleep(1)
            i += 1

            if i > 480:
                logger.error('Canceling task, report generation timeout')
                task.cancel()
                try:
                    await task
                except asyncio.CancelledError:
                    await progress_message.delete()
                    raise WBTimeoutError('WB API timeout after 480 seconds')

        result = await task
        await progress_message.delete()
        return result
    except httpx.HTTPStatusError as e:
        logger.error(f'–û—à–∏–±–∫–∞ –∑–∞–ø—Ä–æ—Å–∞: {e}')
        await progress_message.delete()
        raise InvalidTokenError(f'HTTP error: {e.response.status_code}')
    except (WBTimeoutError, InvalidTokenError, NoDataError):
        raise
    except Exception as e:
        await progress_message.delete()
        raise e


def get_weeks_range(count):
    today = date.today()
    previous_monday = today - timedelta(days=today.weekday()) - timedelta(days=7)
    if today.weekday() == 0:
        previous_monday -= timedelta(days=7)
    weeks_range = []
    for i in range(count):
        week_start = previous_monday - timedelta(weeks=i)
        week_end = week_start + timedelta(days=6)
        weeks_range.append(f'{week_start.strftime("%d.%m.%Y")}-{week_end.strftime("%d.%m.%Y")}')

    return weeks_range


def get_quarters_range():
    today = date.today()
    current_year = today.year
    current_quarter = (today.month - 1) // 3 + 1
    quarters_month = {
        0: '—è–Ω–≤-–º–∞—Ä',
        1: '–∞–ø—Ä-–∏—é–Ω',
        2: '–∏—é–ª-—Å–µ–Ω',
        3: '–æ–∫—Ç-–¥–µ–∫'
    }

    quarters = []
    for year in range(2025, current_year + 1):
        quarters_num = current_quarter if year == current_year else 4
        for quarter in range(quarters_num):
            quarters.append([f'{year}_{quarter}', f'{quarters_month[quarter]} {year}'])

    return quarters


def get_quarters_weeks(year: int, quarter: int):
    today = date.today()
    start_month = quarter * 3 + 1
    first_day = date(year, start_month, 1)
    start_date = first_day + timedelta(days=7) - timedelta(days=first_day.weekday())

    if start_month == 10:  # Q4
        last_day = date(year, 12, 31)
    else:
        last_day = date(year, start_month + 3, 1) - timedelta(days=1)

    if last_day < today:
        end_date = last_day - timedelta(days=last_day.weekday())
    else:
        end_date = today - timedelta(days=today.weekday()) - timedelta(days=7)
        if today.weekday() == 0:
            end_date -= timedelta(days=7)
    weeks_range = []
    current_monday = start_date
    while current_monday <= end_date:
        current_sunday = current_monday + timedelta(days=6)
        weeks_range.append(f'{current_monday.strftime("%d.%m.%Y")}-{current_sunday.strftime("%d.%m.%Y")}')
        current_monday += timedelta(weeks=1)

    return weeks_range


async def orm_add_report(session: AsyncSession, tg_id: int, date_of_week: date, report_path: str, store_id: int):
    obj = Report(
        tg_id=tg_id,
        date_of_week=date_of_week,
        report_path=report_path,
        store_id=store_id,
    )
    session.add(obj)
    await session.commit()


def get_dates_from_str(dates):
    """transform dates DD.MM.YYYY-DD.MM.YYYY to YYYY-MM-DD, YYYY-MM-DD"""
    dates = dates.split('-')
    start_date, end_date = dates[0].split('.'), dates[1].split('.')
    start_date = f'{start_date[2]}-{start_date[1]}-{start_date[0]}'
    end_date = f'{end_date[2]}-{end_date[1]}-{end_date[0]}'
    return start_date, end_date


def change_str_dates(start_date: str, end_date:str, count: int) -> [str, str]:
    start_date = datetime.strptime(start_date, "%Y-%m-%d").date() + timedelta(days=count)
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date() + timedelta(days=count)
    return start_date.isoformat(), end_date.isoformat()


def get_dates_in_range(start: str, end: str) -> List[str]:
    s = datetime.strptime(start, "%Y-%m-%d").date()
    e = datetime.strptime(end,   "%Y-%m-%d").date()
    dates: List[str] = []
    while s <= e:
        dates.append(s.isoformat())
        s += timedelta(days=1)
    return dates


def create_empty_adv_report() -> pd.DataFrame:
    return pd.DataFrame(columns=["–ê—Ä—Ç–∏–∫—É–ª WB", "totalAdjustedSum", "Period"])


async def fetch_product_cards_mapping(token: str) -> Dict:
    logger.info("–ó–∞–≥—Ä—É–∑–∫–∞ –º–∞–ø–ø–∏–Ω–≥–∞ –∫–∞—Ä—Ç–æ—á–µ–∫ —Ç–æ–≤–∞—Ä–∞...")
    url = "https://content-api.wildberries.ru/content/v2/get/cards/list"
    headers = {"Authorization": token, "Content-Type": "application/json"}

    LIMIT = 100          # –Ω–µ —É–≤–µ–ª–∏—á–∏–≤–∞–µ–º
    MAX_RETRIES = 5
    BASE_WAIT = 0.5      # –ø–∞—É–∑–∞ –º–µ–∂–¥—É —É—Å–ø–µ—à–Ω—ã–º–∏ –∑–∞–ø—Ä–æ—Å–∞–º–∏

    payload = {
        "settings": {
            "cursor": {"limit": LIMIT},
            "filter": {"withPhoto": -1},
        }
    }

    mapping: Dict[str, str] = {}

    while True:
        for attempt in range(MAX_RETRIES):
            resp = await ASYNC_CLIENT.post(url, headers=headers, json=payload)

            if resp.status_code == 429:
                retry_after = resp.headers.get("X-Ratelimit-Retry")
                wait = int(retry_after) if retry_after else int(2 ** attempt)
                logger.warning(
                    "429 –æ—Ç content-api (–ø–æ–ø—ã—Ç–∫–∞ %s/%s), –∂–¥—ë–º %s —Å–µ–∫",
                    attempt + 1, MAX_RETRIES, wait
                )
                await asyncio.sleep(wait)
                continue

            resp.raise_for_status()
            break
        else:
            raise RuntimeError("content-api: –ø—Ä–µ–≤—ã—à–µ–Ω–æ —á–∏—Å–ª–æ —Ä–µ—Ç—Ä–∞–µ–≤")

        data = resp.json()
        cards = data.get("cards", [])

        for c in cards:
            key = str(c.get("nmID") or c.get("nmId", "")).strip()
            if key:
                mapping[key] = (c.get("vendorCode") or "").strip()

        cursor = data.get("cursor") or {}

        if not cursor or len(cards) < LIMIT:
            break

        payload["settings"]["cursor"] = {
            "limit": LIMIT,
            "updatedAt": cursor.get("updatedAt"),
            "nmID": cursor.get("nmID"),
        }

        # –æ–±—è–∑–∞—Ç–µ–ª—å–Ω–∞—è –ø–∞—É–∑–∞ –º–µ–∂–¥—É —É—Å–ø–µ—à–Ω—ã–º–∏ –∑–∞–ø—Ä–æ—Å–∞–º–∏
        await asyncio.sleep(BASE_WAIT)

    logger.info("–ú–∞–ø–ø–∏–Ω–≥ –∫–∞—Ä—Ç–æ—á–µ–∫ –∑–∞–≥—Ä—É–∂–µ–Ω: %d –∑–∞–ø–∏—Å–µ–π", len(mapping))
    return mapping


# ------------------ Sales Report ------------------

async def fetch_sales_records_async(date_from: str, date_to: str, token: str) -> list[dict]:
    logger.info("–ù–∞—á–∏–Ω–∞–µ–º –∑–∞–≥—Ä—É–∑–∫—É –æ—Ç—á—ë—Ç–∞ –ø–æ –ø—Ä–æ–¥–∞–∂–∞–º —Å %s –ø–æ %s", date_from, date_to)
    url = "https://statistics-api.wildberries.ru/api/v5/supplier/reportDetailByPeriod"
    headers = {"Authorization": token, "Content-Type": "application/json"}
    records, rrdid = [], 0

    while True:
        resp = await ASYNC_CLIENT.get(
            url,
            headers=headers,
            params={"dateFrom": date_from, "dateTo": date_to, "rrdid": rrdid, "limit": 100000}
        )

        if resp.status_code == 429:
            retry_after = int(resp.headers.get("X-Ratelimit-Retry", 60))
            logger.warning("–ü–æ–ª—É—á–µ–Ω 429 Too Many Requests. –ñ–¥–µ–º %s —Å–µ–∫—É–Ω–¥.", retry_after)
            await asyncio.sleep(retry_after)
            continue  # –ø–æ–≤—Ç–æ—Ä—è–µ–º –∑–∞–ø—Ä–æ—Å –ø–æ—Å–ª–µ –ø–∞—É–∑—ã

        if resp.status_code == 204:
            logger.info("–û—Ç—á–µ—Ç –ø–æ –ø—Ä–æ–¥–∞–∂–∞–º –ø–æ–ª–Ω–æ—Å—Ç—å—é –≤—ã–≥—Ä—É–∂–µ–Ω (204 No Content)")
            break

        resp.raise_for_status()

        chunk = resp.json()
        if not chunk:
            break

        records.extend(chunk)

        last = chunk[-1]
        new_rrdid = last.get("rrd_id") or last.get("rrdid")
        if not new_rrdid or new_rrdid == rrdid:
            break
        rrdid = new_rrdid

        # –ñ–¥–µ–º –ø–µ—Ä–µ–¥ —Å–ª–µ–¥—É—é—â–∏–º –∑–∞–ø—Ä–æ—Å–æ–º –ø–æ –ª–∏–º–∏—Ç—É
        retry_after = resp.headers.get("X-Ratelimit-Retry")
        wait = int(retry_after) if retry_after else 1
        logger.info("–ñ–¥–µ–º %s —Å–µ–∫—É–Ω–¥ –ø–µ—Ä–µ–¥ —Å–ª–µ–¥—É—é—â–∏–º –∑–∞–ø—Ä–æ—Å–æ–º...", wait)
        await asyncio.sleep(wait)

    logger.info("–ó–∞–≥—Ä—É–∑–∫–∞ –æ—Ç—á—ë—Ç–∞ –ø–æ –ø—Ä–æ–¥–∞–∂–∞–º –∑–∞–≤–µ—Ä—à–µ–Ω–∞: %d –∑–∞–ø–∏—Å–µ–π", len(records))
    return records



async def transform_sales_records(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Transform sales records and extract storage data.
    Returns: (sales_df, storage_df) - storage_df contains storage_fee aggregated by nm_id
    """
    empty_sales = pd.DataFrame(columns=[
        "–ê—Ä—Ç–∏–∫—É–ª WB", "SUM –∏–∑ –ö–æ–ª-–≤–æ", "SUM –∏–∑ –°—É–º–º–∞ –ø—Ä–æ–¥–∞–∂", "SUM –∏–∑ –ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É",
        "SUM –∏–∑ –ö–æ–ª-–≤–æ –¥–æ—Å—Ç–∞–≤–æ–∫", "SUM –∏–∑ –°—Ç–æ–∏–º–æ—Å—Ç—å –¥–æ—Å—Ç–∞–≤–∫–∏", "SUM –∏–∑ –®—Ç—Ä–∞—Ñ—ã",
        "SUM –∏–∑ –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–π –ø–ª–∞—Ç–µ–∂", "–£—Ç–∏–ª–∏–∑–∞—Ü–∏—è", "–ü–æ–¥–ø–∏—Å–∫–∞ ¬´–î–∂–µ–º¬ª", "–£–¥–µ—Ä–∂–∞–Ω–∏—è"
    ])
    empty_storage = pd.DataFrame(columns=["nmId", "totalStorageSum"])

    if df.empty:
        return empty_sales, empty_storage
    df = df.copy()
    if "deduction" in df.columns:
        df["deduction"] = pd.to_numeric(df["deduction"], errors="coerce").fillna(0)
    if "bonusTypeName" in df.columns:
        df.rename(columns={"bonusTypeName": "bonus_type_name"}, inplace=True)
    total_util = df.loc[
        df["bonus_type_name"].str.contains("—É—Ç–∏–ª–∏–∑–∞—Ü–∏–∏", case=False, na=False) & (df["deduction"] != 0),
        "deduction"
    ].sum()
    total_jam = df.loc[
        df["bonus_type_name"].str.contains("–¥–∂–µ–º", case=False, na=False) & (df["deduction"] != 0),
        "deduction"
    ].sum()
    df = df[df["nm_id"] != 0]

    if "acceptance" in df.columns:
        acceptance_agg = (
            df.loc[df["acceptance"] != 0]
            .groupby("nm_id", as_index=False)["acceptance"]
            .sum()
            .rename(columns={"acceptance": "–ü—Ä–∏–µ–º–∫–∞"})
        )
    else:
        acceptance_agg = pd.DataFrame(columns=["nm_id", "–ü—Ä–∏–µ–º–∫–∞"])

    sales_df = df[df["doc_type_name"] != "–í–æ–∑–≤—Ä–∞—Ç"].copy()
    returns_df = df[df["doc_type_name"] == "–í–æ–∑–≤—Ä–∞—Ç"].copy()

    sales_agg = sales_df.groupby("nm_id", as_index=False).agg(
        quantity=pd.NamedAgg(
            column="quantity",
            aggfunc=lambda x: x[sales_df.loc[x.index, "doc_type_name"] == "–ü—Ä–æ–¥–∞–∂–∞"].sum()
        ),
        retail_amount=pd.NamedAgg(column="retail_amount", aggfunc="sum"),
        ppvz_for_pay=pd.NamedAgg(column="ppvz_for_pay", aggfunc="sum"),
        delivery_amount=pd.NamedAgg(column="delivery_amount", aggfunc="sum"),
        delivery_rub=pd.NamedAgg(column="delivery_rub", aggfunc="sum"),
        penalty=pd.NamedAgg(column="penalty", aggfunc="sum"),
        additional_payment=pd.NamedAgg(column="additional_payment", aggfunc="sum"),
        cashback_amount=pd.NamedAgg(column="cashback_amount", aggfunc="sum")
    )

    sales_agg.rename(columns={
        "quantity":"SUM –∏–∑ –ö–æ–ª-–≤–æ","retail_amount":"SUM –∏–∑ –°—É–º–º–∞ –ø—Ä–æ–¥–∞–∂",
        "ppvz_for_pay":"SUM –∏–∑ –ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É","delivery_amount":"SUM –∏–∑ –ö–æ–ª-–≤–æ –¥–æ—Å—Ç–∞–≤–æ–∫",
        "delivery_rub":"SUM –∏–∑ –°—Ç–æ–∏–º–æ—Å—Ç—å –¥–æ—Å—Ç–∞–≤–∫–∏","penalty":"SUM –∏–∑ –®—Ç—Ä–∞—Ñ—ã",
        "additional_payment":"SUM –∏–∑ –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–π –ø–ª–∞—Ç–µ–∂"
    }, inplace=True)
    cnt = len(sales_agg)
    sales_agg["–£—Ç–∏–ª–∏–∑–∞—Ü–∏—è"] = round(total_util/cnt,2) if cnt else 0.0
    sales_agg["–ü–æ–¥–ø–∏—Å–∫–∞ ¬´–î–∂–µ–º¬ª"] = round(total_jam/cnt,2) if cnt else 0.0
    returns_agg = returns_df.groupby("nm_id", as_index=False).agg({
        "quantity":"sum","retail_amount":"sum","ppvz_for_pay":"sum"
    })
    returns_agg.rename(columns={
        "quantity":"–í–æ–∑–≤—Ä–∞—Ç—ã (–ö–æ–ª-–≤–æ)","retail_amount":"–í–æ–∑–≤—Ä–∞—Ç—ã (–°—É–º–º–∞ –ø—Ä–æ–¥–∞–∂)",
        "ppvz_for_pay":"–í–æ–∑–≤—Ä–∞—Ç—ã (–ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É)"
    }, inplace=True)
    merged = pd.merge(sales_agg, returns_agg, on="nm_id", how="left")

    merged = pd.merge(merged, acceptance_agg, on="nm_id", how="left")
    merged["–ü—Ä–∏–µ–º–∫–∞"] = merged["–ü—Ä–∏–µ–º–∫–∞"].fillna(0)

    for c in ["–í–æ–∑–≤—Ä–∞—Ç—ã (–ö–æ–ª-–≤–æ)","–í–æ–∑–≤—Ä–∞—Ç—ã (–°—É–º–º–∞ –ø—Ä–æ–¥–∞–∂)","–í–æ–∑–≤—Ä–∞—Ç—ã (–ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É)"]:
        merged[c] = merged[c].fillna(0)
    merged.rename(columns={"nm_id":"–ê—Ä—Ç–∏–∫—É–ª WB"}, inplace=True)

    # Extract storage_fee from sales report (faster than paid_storage API)
    if "storage_fee" in df.columns:
        storage_df = (
            df[df["nm_id"] != 0]
            .groupby("nm_id", as_index=False)["storage_fee"]
            .sum()
            .rename(columns={"nm_id": "nmId", "storage_fee": "totalStorageSum"})
        )
        storage_df["nmId"] = storage_df["nmId"].astype(str).str.upper()
        storage_df["totalStorageSum"] = storage_df["totalStorageSum"].round(2)
        logger.info("–•—Ä–∞–Ω–µ–Ω–∏–µ –∏–∑–≤–ª–µ—á–µ–Ω–æ –∏–∑ –æ—Ç—á—ë—Ç–∞ –ø—Ä–æ–¥–∞–∂: %d –ø–æ–∑–∏—Ü–∏–π", len(storage_df))
    else:
        storage_df = pd.DataFrame(columns=["nmId", "totalStorageSum"])
        logger.warning("storage_fee –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –≤ –æ—Ç—á—ë—Ç–µ –ø—Ä–æ–¥–∞–∂")

    return merged, storage_df


# ------------------ Storage Report (fallback, slow) ------------------

async def get_storage_report(date_from: str, date_to: str, token: str) -> pd.DataFrame:
    logger.info("–ó–∞–ø—Ä–æ—Å –æ—Ç—á—ë—Ç–∞ –ø–æ –ø–ª–∞—Ç–Ω–æ–º—É —Ö—Ä–∞–Ω–µ–Ω–∏—é... %s ‚Äì %s", date_from, date_to)
    base, headers = "https://seller-analytics-api.wildberries.ru/api/v1/paid_storage", {"Authorization":f"Bearer {token}"}
    # create
    for backoff in [5,10,20,40,80]:
        resp = await ASYNC_CLIENT.get(base, headers=headers, params={"dateFrom": date_from, "dateTo": date_to})
        if resp.status_code != 429:
            resp.raise_for_status()
            break
        logger.warning("429 –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ —Ö—Ä–∞–Ω–µ–Ω–∏—è, –∂–¥—É %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["nmId","totalStorageSum"])
    task = resp.json()["data"]["taskId"]
    status_url = f"{base}/tasks/{task}/status"
    # poll
    while True:
        st = await ASYNC_CLIENT.get(status_url, headers=headers)
        if st.status_code == 429:
            await asyncio.sleep(5)
            continue
        st.raise_for_status()
        if st.json()["data"]["status"].lower()=="done":
            break
        await asyncio.sleep(5)
    # download
    dl_url = f"{base}/tasks/{task}/download"
    for backoff in [5,10,20,40,80]:
        dl = await ASYNC_CLIENT.get(dl_url, headers=headers)
        if dl.status_code != 429:
            dl.raise_for_status()
            data = dl.json()
            break
        logger.warning("429 –ø—Ä–∏ —Å–∫–∞—á–∏–≤–∞–Ω–∏–∏ —Ö—Ä–∞–Ω–µ–Ω–∏—è, –∂–¥—É %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["nmId", "totalStorageSum"])
    
    data = dl.json()
    if not data:
        logger.warning("API —Ö—Ä–∞–Ω–µ–Ω–∏—è –≤–µ—Ä–Ω—É–ª –ø—É—Å—Ç–æ–π –º–∞—Å—Å–∏–≤")
        return pd.DataFrame(columns=["nmId", "totalStorageSum"])
    
    df = pd.DataFrame(data)
    
    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞–ª–∏—á–∏–µ –Ω–µ–æ–±—Ö–æ–¥–∏–º—ã—Ö –∫–æ–ª–æ–Ω–æ–∫
    if "nmId" not in df.columns:
        logger.warning("–í –¥–∞–Ω–Ω—ã—Ö —Ö—Ä–∞–Ω–µ–Ω–∏—è –æ—Ç—Å—É—Ç—Å—Ç–≤—É–µ—Ç –∫–æ–ª–æ–Ω–∫–∞ nmId")
        return pd.DataFrame(columns=["nmId", "totalStorageSum"])
    
    if "warehousePrice" in df.columns:
        df["–¶–µ–Ω–∞ —Å–∫–ª–∞–¥–∞"] = pd.to_numeric(df["warehousePrice"], errors="coerce").fillna(0)
    else:
        df["–¶–µ–Ω–∞ —Å–∫–ª–∞–¥–∞"] = 0
    grp = df.groupby("nmId",as_index=False)["–¶–µ–Ω–∞ —Å–∫–ª–∞–¥–∞"].sum().rename(columns={"–¶–µ–Ω–∞ —Å–∫–ª–∞–¥–∞":"totalStorageSum"})
    grp["nmId"] = grp["nmId"].astype(str).str.upper()
    grp["totalStorageSum"] = grp["totalStorageSum"].round(2)
    logger.info("–û—Ç—á—ë—Ç –ø–æ —Ö—Ä–∞–Ω–µ–Ω–∏—é –≥–æ—Ç–æ–≤: %d –ø–æ–∑–∏—Ü–∏–π", len(grp))
    return grp[["nmId", "totalStorageSum"]]


# ------------------ Acceptance report ------------------

async def get_acceptance_report(date_from: str, date_to: str, token: str) -> pd.DataFrame:
    logger.info("–ó–∞–ø—Ä–æ—Å –æ—Ç—á—ë—Ç–∞ –ø–æ –ø–ª–∞—Ç–Ω–æ–π –ø—Ä–∏—ë–º–∫–µ... %s ‚Äì %s", date_from, date_to)
    base, headers = "https://seller-analytics-api.wildberries.ru/api/v1/acceptance_report", {"Authorization":token}
    date_from, date_to = change_str_dates(date_from, date_to, -1)
    # create
    for backoff in [5,10,20,40,80]:
        resp = await ASYNC_CLIENT.get(base, headers=headers, params={"dateFrom": date_from, "dateTo": date_to})
        if resp.status_code != 429:
            resp.raise_for_status()
            break
        logger.warning("429 –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ –ø—Ä–∏—ë–º–∫–∏, –∂–¥—É %s", backoff); await asyncio.sleep(backoff)
    else:
        return pd.DataFrame(columns=["–ê—Ä—Ç–∏–∫—É–ª WB","–ü–ª–∞—Ç–Ω–∞—è –ø—Ä–∏–µ–º–∫–∞"])
    task = resp.json()["data"]["taskId"]
    status_url = f"{base}/tasks/{task}/status"
    while True:
        await asyncio.sleep(5)
        st = await ASYNC_CLIENT.get(status_url, headers=headers)
        st.raise_for_status()
        if st.json()["data"]["status"].lower()=="done":
            break
    dl = await ASYNC_CLIENT.get(f"{base}/tasks/{task}/download", headers=headers)
    dl.raise_for_status()
    data = dl.json()
    if not isinstance(data,list) or not data:
        return pd.DataFrame(columns=["–ê—Ä—Ç–∏–∫—É–ª WB","–ü–ª–∞—Ç–Ω–∞—è –ø—Ä–∏–µ–º–∫–∞"])
    df_ac = pd.DataFrame(data)
    nm_col = next((c for c in df_ac.columns if re.search(r"(?i)nm[_]?id$",c)),None)
    if nm_col is None:
        nm_col = next((c for c in df_ac.columns if re.search(r"(?i)nm.*id",c)),None)
    if nm_col is None:
        return pd.DataFrame(columns=["–ê—Ä—Ç–∏–∫—É–ª WB","–ü–ª–∞—Ç–Ω–∞—è –ø—Ä–∏–µ–º–∫–∞"])
    df_ac["–ê—Ä—Ç–∏–∫—É–ª WB"] = df_ac[nm_col].astype(str).str.upper()
    ac = df_ac.groupby("–ê—Ä—Ç–∏–∫—É–ª WB",as_index=False)["total"].sum().rename(columns={"total":"–ü–ª–∞—Ç–Ω–∞—è –ø—Ä–∏–µ–º–∫–∞"})
    logger.info("–û—Ç—á—ë—Ç –ø–æ –ø—Ä–∏—ë–º–∫–µ –≥–æ—Ç–æ–≤: %d –ø–æ–∑–∏—Ü–∏–π",len(ac))
    return ac


# ------------------ Adds report ------------------

def get_ad_expenses_report(token: str, doc_number: str, period_end: str) -> pd.DataFrame:
    logger.info("–§–æ—Ä–º–∏—Ä–æ–≤–∞–Ω–∏–µ –æ—Ç—á—ë—Ç–∞ –ø–æ —Ä–µ–∫–ª–∞–º–µ, updNum=%s", doc_number)
    if not doc_number:
        return create_empty_adv_report()
    elif ' ' in doc_number:
        upd = set([int(num) for num in doc_number.split()])
    else:
        upd = {int(doc_number)}

    end_date = datetime.strptime(period_end, "%Y-%m-%d").date()
    fr, to = (end_date - timedelta(days=30)).isoformat(), period_end
    period = f"{fr} - {to}"
    headers = {"Authorization": token}

    # –ó–∞–ø—Ä–æ—Å —Å–ø–∏—Å–∫–∞ —Ä–µ–∫–ª–∞–º–Ω—ã—Ö –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤
    upd_list = SYNC_CLIENT.get(
        f"https://advert-api.wildberries.ru/adv/v1/upd?from={fr}&to={to}",
        headers=headers
    )
    upd_list.raise_for_status()

    # –§–∏–ª—å—Ç—Ä–∞—Ü–∏—è –ø–æ –Ω–æ–º–µ—Ä–∞–º –¥–æ–∫—É–º–µ–Ω—Ç–æ–≤
    items = [x for x in upd_list.json() if x.get("updNum") in upd]
    if not items:
        return create_empty_adv_report()

    # –°—É–º–º–∏—Ä—É–µ–º —Ä–∞—Å—Ö–æ–¥—ã –ø–æ –∫–∞–º–ø–∞–Ω–∏—è–º
    summary = {}
    for it in items:
        cid = it.get("advertId") or it.get("id")
        summary[cid] = summary.get(cid, 0.0) + float(it.get("updSum") or 0)

    # –§–æ—Ä–º–∏—Ä—É–µ–º payload –¥–ª—è –∫–∞–∂–¥–æ–≥–æ –Ω–æ–º–µ—Ä–∞ –¥–æ–∫—É–º–µ–Ω—Ç–∞
    payload = []
    for doc_num in upd:
        payload.extend([
            {"id": cid, "updNum": doc_num, "dates": get_dates_in_range(fr, to)}
            for cid in summary
        ])

    # –ó–∞–ø—Ä–æ—Å –¥–µ—Ç–∞–ª—å–Ω–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–∏
    full = SYNC_CLIENT.post(
        "https://advert-api.wildberries.ru/adv/v2/fullstats",
        headers={**headers, "Content-Type": "application/json"},
        content=json.dumps(payload)
    )
    full.raise_for_status()

    data2 = full.json()
    if not isinstance(data2, list):
        return create_empty_adv_report()

    # –ê–≥—Ä–µ–≥–∞—Ü–∏—è –¥–∞–Ω–Ω—ã—Ö –ø–æ —Ç–æ–≤–∞—Ä–∞–º
    agg = {}
    for camp in data2:
        cid = camp.get("advertId") or camp.get("id")
        fact = summary.get(cid, 0.0)
        raw_total = sum(
            float(nm.get("sum") or 0)
            for day in camp.get("days", [])
            for app in day.get("apps", [])
            for nm in app.get("nm", [])
        ) or 0.0
        coef = fact / raw_total if raw_total > 0 else 1.0

        for day in camp.get("days", []):
            for app in day.get("apps", []):
                for nm in app.get("nm", []):
                    nid = nm.get("nmId")
                    name = nm.get("name") or ""
                    val = float(nm.get("sum") or 0) * coef
                    entry = agg.setdefault(nid, {"nmName": name, "totalAdjustedSum": 0.0})
                    entry["totalAdjustedSum"] += val
                    if not entry["nmName"] and name:
                        entry["nmName"] = name

    # –§–æ—Ä–º–∏—Ä—É–µ–º DataFrame
    rows = [
        [str(nid).upper(), round(e["totalAdjustedSum"], 2), period, e["nmName"]]
        for nid, e in agg.items()
    ]
    df_adv = pd.DataFrame(
        rows,
        columns=["–ê—Ä—Ç–∏–∫—É–ª WB", "totalAdjustedSum", "Period", "–ù–∞–∑–≤–∞–Ω–∏–µ —Ç–æ–≤–∞—Ä–∞"]
    )
    logger.info("–û—Ç—á—ë—Ç –ø–æ —Ä–µ–∫–ª–∞–º–µ –≥–æ—Ç–æ–≤: %d –ø–æ–∑–∏—Ü–∏–π", len(df_adv))
    return df_adv


# ------------------ –ì–µ–Ω–µ—Ä–∞—Ü–∏—è –æ—Ç—á—ë—Ç–∞ ------------------

async def generate_report_with_params(progress_state: dict, dates: str, doc_number: str, store_token: str, store_name: str, tg_id: int, store_id: int) -> str:
    """
    Generate report with progress tracking.

    Args:
        progress_state: Dict for updating progress stage (shared with run_with_progress)
        dates: Period in format "DD.MM.YYYY-DD.MM.YYYY"
        doc_number: WB document number(s)
        store_token: WB API token
        store_name: Store name for report header
        tg_id: Telegram user ID
        store_id: Store ID in database
    """
    logger.info("–°—Ç–∞—Ä—Ç –æ—Ç—á—ë—Ç–∞ –¥–ª—è %s: %s", store_name, dates)
    start_date, end_date = get_dates_from_str(dates)

    # Stage 1: Fetch sales data (the longest operation)
    progress_state['stage'] = 'fetch_sales'
    sales_task = fetch_sales_records_async(f"{start_date}T00:00:00", f"{end_date}T23:59:59", store_token)

    # Stage 2: Fetch ads data
    progress_state['stage'] = 'fetch_ads'
    advert_task = asyncio.to_thread(get_ad_expenses_report, store_token, doc_number, end_date)

    # Stage 3: Fetch product cards
    progress_state['stage'] = 'fetch_cards'
    cards_task = fetch_product_cards_mapping(store_token)

    # Stage 4: Fetch storage report (paid_storage API - storage_fee –Ω–µ –≤—Ö–æ–¥–∏—Ç –≤ reportDetailByPeriod)
    progress_state['stage'] = 'fetch_storage'
    storage_task = get_storage_report(start_date, end_date, store_token)

    # Run all fetches in parallel
    raw_records, adv_df, cards, storage_df = await asyncio.gather(
        sales_task, advert_task, cards_task, storage_task
    )

    # Stage 5: Process data
    progress_state['stage'] = 'process'
    df_raw = pd.DataFrame(raw_records)
    sales_df, _ = await transform_sales_records(df_raw)

    # –æ—Ç–∑—ã–≤—ã –∏ –ø—Ä–æ—á–µ–µ
    reviews_agg=pd.DataFrame(columns=["–ê—Ä—Ç–∏–∫—É–ª WB","–°–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤—ã"])
    total_other=0.0
    if "deduction" in df_raw.columns and "bonus_type_name" in df_raw.columns:
        mask_rev=df_raw["bonus_type_name"].str.contains("—Å–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤",case=False,na=False)
        revs=df_raw[mask_rev&(df_raw["deduction"]!=0)].assign(**{
            "–ê—Ä—Ç–∏–∫—É–ª WB": lambda d: d["bonus_type_name"].str.extract(r"—Ç–æ–≤–∞—Ä\s+(\d+)")[0].str.upper()
        })
        if not revs.empty:
            reviews_agg=revs.groupby("–ê—Ä—Ç–∏–∫—É–ª WB",as_index=False)["deduction"].sum().rename(columns={"deduction":"–°–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤—ã"})
        mask_other=(df_raw["deduction"]!=0)&~df_raw["bonus_type_name"].str.contains(
            "–ø–æ–¥–ø–∏—Å–∫–µ ¬´–î–∂–µ–º¬ª|–°–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤|–ü—Ä–æ–¥–≤–∏–∂–µ–Ω–∏–µ|–ê–∫—Ç —É—Ç–∏–ª–∏–∑–∞—Ü–∏–∏ —Ç–æ–≤–∞—Ä–∞",
            case=False,na=False
        )
        total_other += df_raw.loc[mask_other,"deduction"].sum()
    if "penalty" in df_raw.columns:
        total_penalty = df_raw.loc[
            (df_raw["nm_id"] == 0) & (df_raw["penalty"] != 0),
            "penalty"
        ].sum()
        total_other += total_penalty
    if "supplier_oper_name" in df_raw.columns:
        total_additional_payment = df_raw.loc[
            df_raw["supplier_oper_name"].str.contains("—É–¥–µ—Ä–∂–∞–Ω–∏–µ", case=False, na=False) & (df_raw["additional_payment"] != 0),
            "additional_payment"
        ].sum()
        total_other += total_additional_payment

    # –æ–±—ä–µ–¥–∏–Ω—è–µ–º
    for df,col in [(sales_df,"–ê—Ä—Ç–∏–∫—É–ª WB"),(storage_df,"nmId"),(adv_df,"–ê—Ä—Ç–∏–∫—É–ª WB")]:
        df[col]=df[col].astype(str).str.upper()

    # Debug: –ª–æ–≥–∏—Ä—É–µ–º –¥–∞–Ω–Ω—ã–µ –æ —Ö—Ä–∞–Ω–µ–Ω–∏–∏ –ø–µ—Ä–µ–¥ merge
    logger.info("Storage DF –ø–µ—Ä–µ–¥ merge: %d —Å—Ç—Ä–æ–∫, —Å—É–º–º–∞=%.2f", len(storage_df), storage_df["totalStorageSum"].sum() if len(storage_df) > 0 else 0)
    logger.info("Sales DF –∞—Ä—Ç–∏–∫—É–ª—ã (–ø–µ—Ä–≤—ã–µ 5): %s", list(sales_df["–ê—Ä—Ç–∏–∫—É–ª WB"].head()))
    logger.info("Storage DF –∞—Ä—Ç–∏–∫—É–ª—ã (–ø–µ—Ä–≤—ã–µ 5): %s", list(storage_df["nmId"].head()) if len(storage_df) > 0 else [])

    merged=pd.merge(sales_df, storage_df.rename(columns={"nmId":"–ê—Ä—Ç–∏–∫—É–ª WB"})[["–ê—Ä—Ç–∏–∫—É–ª WB","totalStorageSum"]],on="–ê—Ä—Ç–∏–∫—É–ª WB",how="outer")
    merged=pd.merge(merged, adv_df[["–ê—Ä—Ç–∏–∫—É–ª WB","totalAdjustedSum"]],on="–ê—Ä—Ç–∏–∫—É–ª WB",how="outer")
    # acceptance —Ç–µ–ø–µ—Ä—å –±–µ—Ä—ë—Ç—Å—è –∏–∑ sales_df –∫–∞–∫ "–ü—Ä–∏–µ–º–∫–∞" (–Ω–µ –Ω—É–∂–µ–Ω –æ—Ç–¥–µ–ª—å–Ω—ã–π acceptance_report API)
    merged=pd.merge(merged, reviews_agg, on="–ê—Ä—Ç–∏–∫—É–ª WB",how="left")
    merged.fillna(0,inplace=True)
    merged.sort_values("–ê—Ä—Ç–∏–∫—É–ª WB",inplace=True)

    merged["vendorCode"] = merged["–ê—Ä—Ç–∏–∫—É–ª WB"].map(cards).fillna('–ù–µ—Ä–∞—Å–ø–æ–∑–Ω–∞–Ω–Ω—ã–π –∞—Ä—Ç–∏–∫—É–ª')

    # –ü—Ä–æ—á–∏–µ —É–¥–µ—Ä–∂–∞–Ω–∏—è
    n=len(merged)
    per_item=round(total_other/n,2) if n else 0.0
    merged["–ü—Ä–æ—á–∏–µ —É–¥–µ—Ä–∂–∞–Ω–∏—è"]=per_item

    # –ü–µ—Ä–µ–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ –∏ –ø–æ—Ä—è–¥–æ–∫
    merged.rename(columns={
        "vendorCode":"–ê—Ä—Ç–∏–∫—É–ª –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞",
        "SUM –∏–∑ –ö–æ–ª-–≤–æ":"–ö–æ–ª-–≤–æ –ø—Ä–æ–¥–∞–∂",
        "SUM –∏–∑ –°—É–º–º–∞ –ø—Ä–æ–¥–∞–∂":"–û–±—â–∞—è –≤—ã—Ä—É—á–∫–∞",
        "SUM –∏–∑ –ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É":"–ö –ü–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é",
        "SUM –∏–∑ –ö–æ–ª-–≤–æ –¥–æ—Å—Ç–∞–≤–æ–∫":"–õ–æ–≥–∏—Å—Ç–∏–∫–∞, —à—Ç",
        "SUM –∏–∑ –°—Ç–æ–∏–º–æ—Å—Ç—å –¥–æ—Å—Ç–∞–≤–∫–∏":"–õ–æ–≥–∏—Å—Ç–∏–∫–∞, —Ä—É–±",
        "SUM –∏–∑ –®—Ç—Ä–∞—Ñ—ã":"–®—Ç—Ä–∞—Ñ—ã",
        "SUM –∏–∑ –î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–π –ø–ª–∞—Ç–µ–∂":"–î–æ–ø–ª–∞—Ç—ã",
        "–í–æ–∑–≤—Ä–∞—Ç—ã (–ö –ø–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é –ø—Ä–æ–¥–∞–≤—Ü—É)":"–í–æ–∑–≤—Ä–∞—Ç—ã",
        "totalStorageSum":"–•—Ä–∞–Ω–µ–Ω–∏–µ",
        "totalAdjustedSum":"–í–ë.–ü—Ä–æ–¥–≤–∏–∂–µ–Ω–∏–µ",
        "cashback_amount":"–ë–∞–ª–ª—ã –ø—Ä–æ–≥—Ä–∞–º–º—ã –ª–æ—è–ª—å–Ω–æ—Å—Ç–∏"
    },inplace=True)

    final_cols=[
        "–ê—Ä—Ç–∏–∫—É–ª WB","–ê—Ä—Ç–∏–∫—É–ª –ø–æ—Å—Ç–∞–≤—â–∏–∫–∞","–ö–æ–ª-–≤–æ –ø—Ä–æ–¥–∞–∂","–û–±—â–∞—è –≤—ã—Ä—É—á–∫–∞",
        "–ö –ü–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é","–õ–æ–≥–∏—Å—Ç–∏–∫–∞, —à—Ç","–õ–æ–≥–∏—Å—Ç–∏–∫–∞, —Ä—É–±","–®—Ç—Ä–∞—Ñ—ã","–î–æ–ø–ª–∞—Ç—ã",
        "–í–æ–∑–≤—Ä–∞—Ç—ã","–•—Ä–∞–Ω–µ–Ω–∏–µ","–í–ë.–ü—Ä–æ–¥–≤–∏–∂–µ–Ω–∏–µ","–ü–æ–¥–ø–∏—Å–∫–∞ ¬´–î–∂–µ–º¬ª",
        "–ü—Ä–∏–µ–º–∫–∞","–£—Ç–∏–ª–∏–∑–∞—Ü–∏—è","–°–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤—ã","–ü—Ä–æ—á–∏–µ —É–¥–µ—Ä–∂–∞–Ω–∏—è","–ë–∞–ª–ª—ã –ø—Ä–æ–≥—Ä–∞–º–º—ã –ª–æ—è–ª—å–Ω–æ—Å—Ç–∏"
    ]
    final_df=merged[final_cols]

    # –î–æ–±–∞–≤–ª—è–µ–º —Å—Ç–æ–ª–±–µ—Ü "–ù–∞ —Ä–∞—Å—á–µ—Ç–Ω—ã–π —Å—á–µ—Ç"
    final_df["–ù–∞ —Ä–∞—Å—á–µ—Ç–Ω—ã–π —Å—á–µ—Ç"] = (
        final_df["–ö –ü–µ—Ä–µ—á–∏—Å–ª–µ–Ω–∏—é"]
        - final_df["–õ–æ–≥–∏—Å—Ç–∏–∫–∞, —Ä—É–±"]
        - final_df["–®—Ç—Ä–∞—Ñ—ã"]
        + final_df["–î–æ–ø–ª–∞—Ç—ã"]
        - final_df["–í–æ–∑–≤—Ä–∞—Ç—ã"]
        - final_df["–•—Ä–∞–Ω–µ–Ω–∏–µ"]
        - final_df["–í–ë.–ü—Ä–æ–¥–≤–∏–∂–µ–Ω–∏–µ"]
        - final_df["–ü–æ–¥–ø–∏—Å–∫–∞ ¬´–î–∂–µ–º¬ª"]
        - final_df["–ü—Ä–∏–µ–º–∫–∞"]
        - final_df["–£—Ç–∏–ª–∏–∑–∞—Ü–∏—è"]
        - final_df["–°–ø–∏—Å–∞–Ω–∏–µ –∑–∞ –æ—Ç–∑—ã–≤—ã"]
        - final_df["–ü—Ä–æ—á–∏–µ —É–¥–µ—Ä–∂–∞–Ω–∏—è"]
        - final_df["–ë–∞–ª–ª—ã –ø—Ä–æ–≥—Ä–∞–º–º—ã –ª–æ—è–ª—å–Ω–æ—Å—Ç–∏"]
    )

    # Stage 5: Create Excel file
    progress_state['stage'] = 'create_excel'

    yellow=PatternFill(fill_type="solid",start_color="FFFF00",end_color="FFFF00")

    output_folder = Path(os.getenv('DATA_ROOT')) / 'reports' / str(tg_id) / str(store_id)
    output_folder.mkdir(parents=True, exist_ok=True)
    path = output_folder / f'report{start_date}.xlsx'

    with pd.ExcelWriter(path,engine="openpyxl") as writer:
        final_df.to_excel(writer,index=False,startrow=2)
        ws=writer.sheets["Sheet1"]
        ws.cell(row=1,column=1,value=f"–ú–∞–≥–∞–∑–∏–Ω: {store_name}")
        ws.cell(row=2,column=1,value=f"–ü–µ—Ä–∏–æ–¥: {start_date} ‚Äì {end_date}")
        for cell in ws[3]:
            cell.font=Font(bold=True)
        for col in ws.columns:
            length=max(len(str(c.value)) for c in col)
            ws.column_dimensions[col[0].column_letter].width=length+2
        summary=ws.max_row+1
        ws.cell(row=summary,column=1,value="–ò—Ç–æ–≥–æ").font=Font(bold=True)
        for idx in range(3,len(final_cols)+2):
            letter=get_column_letter(idx)
            c=ws.cell(row=summary,column=idx,value=f"=SUM({letter}4:{letter}{summary-1})")
            c.font=Font(color="FF0000"); c.fill=yellow

    logger.info(f'–ò—Ç–æ–≥–æ–≤—ã–π –æ—Ç—á—ë—Ç —Å–æ—Ö—Ä–∞–Ω—ë–Ω –≤ "{path}"')
    return str(path)